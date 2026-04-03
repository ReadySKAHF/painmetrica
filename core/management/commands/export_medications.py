import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from django.core.management.base import BaseCommand

from medications.models import Medication


THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', fgColor='2E67F6')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
WRAP = Alignment(vertical='top', wrap_text=True)


class Command(BaseCommand):
    help = 'Экспортирует все лекарства в Excel файл'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            default='medications_export.xlsx',
            help='Имя выходного файла (по умолчанию: medications_export.xlsx)',
        )

    def handle(self, *args, **options):
        output = options['output']

        medications = Medication.objects.all().order_by('id')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Лекарства'

        # ── Заголовки ────────────────────────────────────────────────────────
        headers = [
            'ID',
            'Название лекарства',
            'Тип лекарства',
            'Схема применения',
            'Побочные эффекты',
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER

        ws.row_dimensions[1].height = 22

        # ── Данные ───────────────────────────────────────────────────────────
        for row, med in enumerate(medications, 2):
            values = [
                med.id,
                med.name,
                med.medication_type or '—',
                med.prescription_scheme or '—',
                med.side_effects or '—',
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = BORDER
                cell.alignment = CENTER if col == 1 else WRAP

        # ── Ширина столбцов ───────────────────────────────────────────────────
        col_widths = [8, 30, 20, 50, 50]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        wb.save(output)
        self.stdout.write(self.style.SUCCESS(f'Файл сохранён: {output} ({medications.count()} лекарств)'))
